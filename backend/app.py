from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
import os
from pathlib import Path
from zoneinfo import ZoneInfo
from hmac import compare_digest
import logging
from datetime import datetime, timezone
from dateutil import parser
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill
import re
from random import choice
from flask import request
from dotenv import load_dotenv
if os.getenv("FLASK_ENV", "production") != "production":
    load_dotenv()

load_dotenv()

app = Flask(__name__)

ALLOWED_ORIGINS = [
    o.strip() for o in os.getenv("FRONTEND_ORIGINS", "").split(",") if o.strip()
] or ["*"]  

CORS(
    app,
    resources={r"/*": {"origins": ALLOWED_ORIGINS if "*" not in ALLOWED_ORIGINS else "*"}},
    supports_credentials=False,
    methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)


@app.after_request
def add_cors_headers(resp):
    origin = request.headers.get("Origin")
    allow_all = "*" in ALLOWED_ORIGINS
    if origin and (allow_all or origin in ALLOWED_ORIGINS):
        resp.headers["Access-Control-Allow-Origin"] = origin
        resp.headers["Vary"] = "Origin"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
        resp.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        resp.headers["Access-Control-Max-Age"] = "86400"
    return resp

# database
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URI")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Keep the connection pool healthy on hosts that close idle conns
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,    # tests connections before using
    "pool_recycle": 300,      # recycle connections every 5 mins
}

# logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(levelname)s: %(message)s')

db = SQLAlchemy(app)

# storage
UPLOAD_FOLDER = Path('uploaded_templates')
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)

# app constants
TIMEZONE = ZoneInfo("Europe/Dublin")
ALLOWED_PRINT_HOURS = {16, 17, 18}
DASH_PATTERN = r"[-–—]"

# Define a Worker model
class Worker(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    roles = db.Column(db.JSON, nullable=False)  # Roles as a JSON column
    availability = db.Column(db.JSON, nullable=False)  # Availability as a JSON column

    def __repr__(self):
        return f"<Worker {self.name}>"

# API endpoint to get all workers
@app.route("/workers", methods=["GET"])
def get_all_workers():
    try:
        workers = Worker.query.all()
        workers_list = [
            {
                "id": worker.id,
                "name": worker.name,
                "roles": worker.roles,
                "availability": worker.availability
            }
            for worker in workers
        ]
        logging.info("Fetched all workers successfully.")
        return jsonify({"workers": workers_list}), 200
    except Exception as e:
        logging.error(f"Error fetching workers: {e}")
        return jsonify({"error": str(e)}), 500
    
@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    try:
        file = request.files['file']
        if not file:
            return jsonify({'error': 'No file provided'}), 400

        filename = file.filename
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        return jsonify({'message': f'File {filename} uploaded successfully'}), 200
    except Exception as e:
        logging.error(f"Error uploading file: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/upload-worker-availability', methods=['POST'])
def upload_worker_availability():
    try:
        file = request.files['file']
        if not file:
            return jsonify({'error': 'No file provided'}), 400

        workbook = openpyxl.load_workbook(file)

        # --- Helper: robust time-range parser (handles -, – , —, spaces, 24h and 12h AM/PM) ---
        DASH_PATTERN = r"[-–—]"  # hyphen, en dash, em dash

        def parse_time_range(cell_val):
            """
            Accepts strings like '08:00 - 16:00', '08:00–16:00', '8:00 AM — 4:30 PM'.
            Returns (start_time, end_time) as datetime.time.
            """
            s = str(cell_val).strip()

            m = re.search(rf"(\d{{1,2}}:\d{{2}})\s*{DASH_PATTERN}\s*(\d{{1,2}}:\d{{2}})", s)
            if m:
                t1 = datetime.strptime(m.group(1), "%H:%M").time()
                t2 = datetime.strptime(m.group(2), "%H:%M").time()
                return t1, t2

            m = re.search(rf"(\d{{1,2}}:\d{{2}}\s*[APap][Mm])\s*{DASH_PATTERN}\s*(\d{{1,2}}:\d{{2}}\s*[APap][Mm])", s)
            if m:
                t1 = datetime.strptime(m.group(1).upper(), "%I:%M %p").time()
                t2 = datetime.strptime(m.group(2).upper(), "%I:%M %p").time()
                return t1, t2

            raise ValueError(f"Unrecognized time range: {cell_val!r}")

        all_results = []  # collects a summary across all sheets
        updated_count = 0  # counter of DB updates

        # Process ALL worksheets in the file
        for sheet in workbook.worksheets:
            logging.info(f"🔎 Processing sheet: {sheet.title}")

            # Step 1: Extract date from B22 area on this sheet
            target_date = None
            for row in sheet.iter_rows(min_row=22, max_row=22):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and re.search(r"\d{2}/\d{2}/\d{4}", cell.value):
                        match = re.search(r"\d{2}/\d{2}/\d{4}", cell.value)
                        if match:
                            target_date = datetime.strptime(match.group(), "%d/%m/%Y")
                            break
                if target_date:
                    break

            if not target_date:
                logging.warning(f"⚠️ Skipping sheet '{sheet.title}' — no date found on row 22.")
                continue  # move to next sheet

            # Step 2: Collect names and times from row 24 down on this sheet
            for i, row in enumerate(sheet.iter_rows(min_row=24), start=24):
                name_cell = row[1] if len(row) > 1 else None  # Column B
                time_cell = None

                # Check columns D, E, F for a single-cell range first
                for idx in [3, 4, 5]:
                    if len(row) > idx and row[idx].value:
                        time_cell = row[idx]
                        break

                logging.info(f"[{sheet.title}] Row {i} -> name: {name_cell.value if name_cell else 'None'}, time: {time_cell.value if time_cell else 'None'}")

                if not (name_cell and name_cell.value):
                    continue

                worker_name = str(name_cell.value).strip()

                # Try single-cell time range first
                start_t = end_t = None
                if time_cell and time_cell.value:
                    try:
                        start_t, end_t = parse_time_range(time_cell.value)
                    except Exception as parse_err:
                        logging.debug(f"[{sheet.title}] Single-cell time parse failed at row {i}: {parse_err}")

                # Fallback: if range not found in one cell, try separate start/end in D and E
                if (start_t is None or end_t is None):
                    start_cell = row[3] if len(row) > 3 else None  # col D
                    end_cell   = row[4] if len(row) > 4 else None  # col E

                    def to_time(val):
                        """
                        Convert an Excel cell value into datetime.time if possible.
                        Handles datetime, time, 'HH:MM', and 'HH:MM AM/PM'.
                        """
                        if val is None:
                            return None
                        if isinstance(val, datetime):
                            return val.time()
                        try:
                            from datetime import time as _time
                            if isinstance(val, _time):
                                return val
                        except Exception:
                            pass
                        if isinstance(val, str):
                            s = val.strip().upper()
                            try:
                                return datetime.strptime(s, "%H:%M").time()
                            except ValueError:
                                pass
                            try:
                                return datetime.strptime(s, "%I:%M %p").time()
                            except ValueError:
                                return None
                        return None

                    if start_cell and end_cell:
                        start_t = to_time(start_cell.value)
                        end_t = to_time(end_cell.value)

                if not (start_t and end_t):
                    # Nothing parseable on this row; continue to next row
                    continue

                # Record for response logging (keeps your existing behavior)
                time_range_display = f"{start_t.strftime('%H:%M')} - {end_t.strftime('%H:%M')}"
                all_results.append({"sheet": sheet.title, "name": worker_name, "time": time_range_display, "date": target_date.strftime("%Y-%m-%d")})

                # Update each worker's availability in the database (per-sheet date)
                existing_worker = Worker.query.filter_by(name=worker_name).first()
                if existing_worker:
                    try:
                        # Build datetimes on sheet's target_date using Europe/London timezone
                        start_datetime = datetime.combine(target_date.date(), start_t).replace(tzinfo=ZoneInfo("Europe/London"))
                        end_datetime   = datetime.combine(target_date.date(), end_t).replace(tzinfo=ZoneInfo("Europe/London"))

                        new_availability = {
                            "start": start_datetime.isoformat(),
                            "end": end_datetime.isoformat(),
                            "late": False
                        }

                        # Remove existing availability for this date (if any), then append
                        updated_availability = [
                            a for a in existing_worker.availability
                            if parser.parse(a["start"]).date() != target_date.date()
                        ]
                        updated_availability.append(new_availability)
                        existing_worker.availability = updated_availability

                        updated_count += 1
                    except Exception as parse_err:
                        logging.warning(f" Could not save time for {worker_name} (sheet '{sheet.title}', row {i}): {parse_err}")
                else:
                    logging.warning(f" Worker not found in DB: {worker_name}")

        # Commit once after processing all sheets (reduces I/O)
        db.session.commit()

        # Log the parsed availability (across all sheets)
        logging.info(" Parsed worker availability from Excel (all sheets):")
        for entry in all_results:
            logging.info(f"[{entry['sheet']}] {entry['date']} — {entry['name']} - {entry['time']}")
        logging.info(f"Total availability updates: {updated_count}")

        # Build response summary by date/sheet
        return jsonify({
            "updates": updated_count,
            "entries": all_results
        }), 200

    except Exception as e:
        logging.error(f" Error parsing availability upload: {e}")
        return jsonify({'error': str(e)}), 500


    
@app.route('/list-templates', methods=['GET'])
def list_templates():
    try:
        files = os.listdir(UPLOAD_FOLDER)
        logging.debug(f"Templates found: {files}")
        return jsonify({'templates': files}), 200
    except Exception as e:
        logging.error(f"Error listing templates: {e}")
        return jsonify({'error': str(e)}), 500

# API endpoint to generate the schedule and save to Excel
@app.route('/generate-schedule', methods=['POST'])
def generate_schedule():
    try:
        selected_file = request.json.get('template')
        selected_date_str = request.json.get('date')  # Get selected date from request

        ica_morning_count = request.json.get("ica_morning_count", 4)
        ica_afternoon_count = request.json.get("ica_afternoon_count", 4)

        # Clamp values to stay between 2 and 4
        ica_morning_count = max(2, min(4, int(ica_morning_count)))
        ica_afternoon_count = max(2, min(4, int(ica_afternoon_count)))

        # option to extend non-ICA printing to 4pm, 5pm, or 6pm
        print_until_hour = int(request.json.get("print_until_hour", 16))
        if print_until_hour not in (16, 17, 18):
            print_until_hour = 16

        # Reset all stateful variables to prevent carryover issues
        valid_roles = {}
        used_workers = set()
        afternoon_valid_roles = {}  
        afternoon_used_workers = set()
        course_workers = []  # Ensures fresh assignment
        morning_assignments = {}  # Ensures morning roles are properly tracked

        # Validate input
        if not selected_file:
            return jsonify({'error': 'Template is required'}), 400
        if not selected_date_str:
            return jsonify({'error': 'Date is required'}), 400

        try:
            selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)  # ⬅️ Convert to datetime
        if not selected_file:
            return jsonify({'error': 'No template selected'}), 400

        filepath = os.path.join(UPLOAD_FOLDER, selected_file)
        if not os.path.exists(filepath):
            return jsonify({'error': 'Selected template not found'}), 404

        # Fetch all workers from the database
        workers = Worker.query.all()

        # Separate workers into available and late-shift workers based on the selected date
        in_today_workers = []
        late_shift_workers = []
        untrained_workers = []

        for worker in workers:
            for availability in worker.availability:
                try:
                    start = parser.parse(availability['start']).astimezone(timezone.utc)
                    end = parser.parse(availability['end']).astimezone(timezone.utc)

                    if start.date() <= selected_date.date() <= end.date():
                        if bool(availability.get("late", False)):
                            late_shift_workers.append(worker)
                        else:
                            in_today_workers.append(worker)

                        # Check if worker is untrained
                        if not any(role in worker.roles for role in ["KITUP", "AATT", "MT", "ICA"]):
                            untrained_workers.append(worker)  # Store them separately
                        break  # Stop checking once availability is confirmed
                except Exception as e:
                    logging.error(f"Error parsing availability for worker {worker.name}: {e}")


        # Define role-to-training mapping
        role_to_training = {
            'KITUP': ['Host', 'Dekit', 'Kit Up 1', 'Kit Up 2', 'Kit Up 3', 'Clip In 1', 'Clip In 2'],
            'AATT': ['TREE TREK 1', 'TREE TREK 2', 'Course Support 1', 'Course Support 2', 'Zip Top 1', 'Zip Top 2', 'Zip Ground', 'rotate to course 1'],
            'MT': ['Mini Trek'],
            'ICA': ['ICA 1', 'ICA 2', 'ICA 3', 'ICA 4']
        }

        # Dynamically map roles based on the Excel file
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        role_to_column = {}
        header_row = 1  

        for col in range(1, sheet.max_column + 1):
            role = sheet.cell(row=header_row, column=col).value
            if role:  # Skip empty cells
                role_to_column[role.strip()] = col

        ica_roles_morning = [f"ICA {i}" for i in range(1, ica_morning_count + 1)]

        prioritized_roles_morning = (
            ica_roles_morning +
            [
                'Mini Trek', 'Course Support 2', 'Zip Top 1', 'Zip Top 2', 'Zip Ground', 'rotate to course 1',
                'TREE TREK 1', 'TREE TREK 2',
                'Clip In 1', 'Clip In 2', 'Kit Up 3', 'Kit Up 2', 'Kit Up 1',
            ]
        )

        # Only add 'Course Support 1' if it exists in the Excel file
        if "Course Support 1" in role_to_column:
            prioritized_roles_morning.insert(5, "Course Support 1")  # Insert at the correct position

        valid_roles = {}
        used_workers = set()

        def get_eligible_workers(role):
            eligible = [
                worker for worker in (in_today_workers + late_shift_workers)
                if worker.name not in used_workers
            ]

            logging.debug(f"🔍 Role: {role} | Eligible workers before filtering: {[w.name for w in eligible]}")

            # KITUP Roles
            if role in role_to_training['KITUP']:
                return [
                    worker for worker in (in_today_workers + late_shift_workers)
                    if worker.name not in used_workers
                    and 'KITUP' in worker.roles  # Must be trained in KITUP
                ]

            # AATT Roles
            elif role in role_to_training['AATT']:
                return [
                    worker for worker in (in_today_workers + late_shift_workers)
                    if worker.name not in used_workers
                    and 'AATT' in worker.roles  # Must be trained in AATT
                ]

            # Mini Trek - Only early workers can be assigned
            elif role in role_to_training['MT']:
                return [
                    worker for worker in in_today_workers  # Only early workers
                    if worker.name not in used_workers
                    and 'MT' in worker.roles  # Must be trained in Mini Trek
                ]

            # ICA - Only early workers can be assigned
            elif role in role_to_training['ICA']:
                return [
                    worker for worker in in_today_workers  # Only early workers
                    if worker.name not in used_workers
                    and 'ICA' in worker.roles  # Must be trained in ICA
                ]

            # RESTRICT LATE-SHIFT WORKERS from Course Support 2, Zip Top 1, Zip Top 2, and Zip Ground
            elif role in ["Course Support 2", "Zip Top 1", "Zip Top 2", "Zip Ground"]:
                return [
                    worker for worker in in_today_workers  # ONLY early workers allowed
                    if worker.name not in used_workers
                    and 'AATT' in worker.roles  # Must be trained in AATT
                ]

            return []

        morning_assignments = {}  # Dictionary to store the morning role assignments

        # Assign workers to the first time slot (9:00-9:30)
        for role in prioritized_roles_morning:
            if role in role_to_column:
                eligible_workers = get_eligible_workers(role)
                logging.debug(f"Checking role: {role} | Eligible workers: {[w.name for w in eligible_workers]}")

                # Exclude late-shift workers for specific roles at 9:00 AM
                if role in ["Course Support 2", "Zip Top 1", "Zip Top 2", "Zip Ground"]:
                    eligible_workers = [worker for worker in eligible_workers if worker not in late_shift_workers]

                if eligible_workers:
                    selected_worker = choice(eligible_workers)

                    if selected_worker.name in used_workers:
                        logging.warning(f"Worker {selected_worker.name} was already marked as used before being assigned to {role}!")

                    logging.debug(f"Assigning {selected_worker.name} to {role} from {len(eligible_workers)} options")

                    valid_roles[role] = selected_worker.name
                    used_workers.add(selected_worker.name)

                else:
                    logging.warning(f"No eligible workers found for {role}")

        # Ensure unassigned KITUP-trained workers are placed in Kit Up roles BEFORE Clip In or other roles
        kitup_roles_priority = ['Kit Up 1', 'Kit Up 2']  # Highest priority
        kitup_roles_secondary = ['Kit Up 3', 'Clip In 1', 'Clip In 2']  # Lower priority

        # Fix: Sort KITUP workers by experience (if needed) or randomize the list
        unassigned_kitup_workers = [
            worker for worker in in_today_workers if worker.name not in used_workers and 'KITUP' in worker.roles
        ]

        # Fix: Ensure no KITUP worker is left unassigned
        if not unassigned_kitup_workers:
            logging.warning("No unassigned KITUP-trained workers available!")

        # Assign to Kit Up 1 & Kit Up 2 first
        for role in kitup_roles_priority:
            if role in role_to_column and role not in valid_roles and unassigned_kitup_workers:
                selected_worker = unassigned_kitup_workers.pop(0)  # Assign first available KITUP worker
                valid_roles[role] = selected_worker.name
                used_workers.add(selected_worker.name)
                logging.debug(f"Assigning {selected_worker.name} to {role} (KITUP priority role)")

        # Assign to other Kit Up/Clip In roles after that
        for role in kitup_roles_secondary:
            if role in role_to_column and role not in valid_roles and unassigned_kitup_workers:
                selected_worker = unassigned_kitup_workers.pop(0)  # Assign first available KITUP worker
                valid_roles[role] = selected_worker.name
                used_workers.add(selected_worker.name)
                logging.debug(f"Assigning {selected_worker.name} to {role} (Secondary KITUP role)")

        # Fix: Log if any KITUP-trained workers are left unassigned (shouldn’t happen)
        if unassigned_kitup_workers:
            logging.warning(f"These KITUP-trained workers were NOT assigned but should be: {[w.name for w in unassigned_kitup_workers]}")


        # Identify untrained workers (people without KITUP, AATT, MT, or ICA)
        untrained_workers = [
            worker for worker in in_today_workers if worker.name not in used_workers
        ]

        # Ensure all Kit Up roles are filled FIRST
        kitup_roles_priority = ['Kit Up 1', 'Kit Up 2']  # Highest priority
        kitup_roles_secondary = ['Kit Up 3', 'Clip In 1', 'Clip In 2']  # Lower priority

        unassigned_kitup_workers = [
            worker for worker in in_today_workers if worker.name not in used_workers and 'KITUP' in worker.roles
        ]

        # Assign to Kit Up 1 & Kit Up 2 first
        for role in kitup_roles_priority:
            if role in role_to_column and role not in valid_roles and unassigned_kitup_workers:
                selected_worker = unassigned_kitup_workers.pop(0)  # Assign first available KITUP worker
                valid_roles[role] = selected_worker.name
                used_workers.add(selected_worker.name)
                logging.debug(f"Assigning {selected_worker.name} to {role} (KITUP priority role)")

        # Assign to other Kit Up/Clip In roles after that
        for role in kitup_roles_secondary:
            if role in role_to_column and role not in valid_roles and unassigned_kitup_workers:
                selected_worker = unassigned_kitup_workers.pop(0)  # Assign first available KITUP worker
                valid_roles[role] = selected_worker.name
                used_workers.add(selected_worker.name)
                logging.debug(f"Assigning {selected_worker.name} to {role} (Secondary KITUP role)")

        # Now, Assign Host & Dekit AFTER all Kit Up roles are filled
        unassigned_workers = [
            worker for worker in in_today_workers if worker.name not in used_workers
        ]

        assigned_host_dekit = set()  # Track assigned workers for Host & Dekit

        for role in ["Host", "Dekit"]:
            if role in role_to_column and role not in valid_roles:  # Only assign if still empty
                available_untrained = [
                    worker for worker in unassigned_workers
                    if worker.name not in assigned_host_dekit  # Ensure a different worker is assigned
                ]

                if available_untrained:
                    selected_worker = choice(available_untrained)
                    valid_roles[role] = selected_worker.name  # Assign worker
                    used_workers.add(selected_worker.name)  # Mark them as used
                    assigned_host_dekit.add(selected_worker.name)  # Track to avoid duplicate assignment

                    logging.debug(f"Assigning {selected_worker.name} to {role} (Untrained Worker)")

        # Ensure Host & Dekit are printed for all morning time slots (9:00 AM - 12:45 PM)
        for slot_row in range(2, 9):  # Covers 9:00–9:30 to 12:45
            for role in ["Host", "Dekit"]:
                column = role_to_column.get(role)
                if column:
                    sheet.cell(row=slot_row, column=column).value = valid_roles.get(role, "")

        # Track morning assignments properly (store all roles)
        for role, worker in valid_roles.items():
            if worker not in morning_assignments:
                morning_assignments[worker] = []  # Initialize list if not present
            morning_assignments[worker].append(role)  # Store all roles they worked in the morning
        
        # Check if any worker was ignored for assignment
        assigned_workers = set(valid_roles.values())
        unassigned_workers = [worker.name for worker in in_today_workers + late_shift_workers if worker.name not in assigned_workers]

        if unassigned_workers:
            logging.warning(f"Workers NOT assigned in the morning (shouldn't happen): {', '.join(unassigned_workers)}")

        logging.info("\n======= MORNING ASSIGNMENTS CHECK =======")
        for role, worker in valid_roles.items():
            logging.info(f"{role} -> {worker}")
        logging.info("========================================")

        # Fallback: Force assign Host and Dekit if still unassigned and spares exist
        for role in ["Host", "Dekit"]:
            if role in role_to_column and role not in valid_roles:
                available_spares = [
                    worker for worker in in_today_workers + late_shift_workers
                    if worker.name not in used_workers
                ]

                if available_spares:
                    selected_worker = choice(available_spares)
                    valid_roles[role] = selected_worker.name
                    used_workers.add(selected_worker.name)
                    logging.warning(f"Fallback assigning {selected_worker.name} to {role} due to earlier miss.")

        # Recalculate spare workers AFTER fallback assignment
        morning_spare_workers = [
            worker.name for worker in in_today_workers + late_shift_workers
            if worker.name not in used_workers
        ]


        # Log morning spare workers clearly
        if morning_spare_workers:
            logging.info(f"Morning Spare Workers ({len(morning_spare_workers)}): {', '.join(morning_spare_workers)}")
        else:
            logging.info("No Morning Spare Workers found.")
        
        
        # Write the assignments for the first time slot (9:00-9:30) to the Excel file
        for role, column in role_to_column.items():
            if role == "Course Support 1" and role not in role_to_column:
                continue  # Skip if it doesn’t exist

            assigned_worker = valid_roles.get(role)
            if assigned_worker:
                sheet.cell(row=2, column=column).value = assigned_worker

        # Fill Shed (Host, Dekit, Kit Up 1), Tree Trek, Mini Trek, and ICA roles for all time slots till lunch
        lunch_slots = [3, 4, 5, 6, 7, 8]  # Rows corresponding to 9:30, 10:00, ..., 12:00-12:45
        for slot_row in lunch_slots:
            for role in ['Host', 'Dekit', 'Kit Up 1', 'TREE TREK 1', 'TREE TREK 2', 'Mini Trek', 'ICA 1', 'ICA 2', 'ICA 3', 'ICA 4']:
                column = role_to_column.get(role)
                if column:
                    sheet.cell(row=slot_row, column=column).value = valid_roles.get(role)

        # Handle Kit Up 2, Kit Up 3, Clip In 1, and Clip In 2
        for slot_row in lunch_slots:
            if slot_row == 5:  # 10:30 row
                # Swap Kit Up 2 with Clip In 1
                temp_kit_up_2 = valid_roles.get('Kit Up 2')
                temp_kit_up_3 = valid_roles.get('Kit Up 3')
                valid_roles['Kit Up 2'], valid_roles['Clip In 1'] = valid_roles.get('Clip In 1'), temp_kit_up_2
                valid_roles['Kit Up 3'], valid_roles['Clip In 2'] = valid_roles.get('Clip In 2'), temp_kit_up_3

            # Assign updated roles for Kit Up 2, Kit Up 3, Clip In 1, and Clip In 2
            for role in ['Kit Up 2', 'Kit Up 3', 'Clip In 1', 'Clip In 2']:
                column = role_to_column.get(role)
                if column:
                    sheet.cell(row=slot_row, column=column).value = valid_roles.get(role)

        # Handle Course role rotations till lunch
        course_roles = [
            'Course Support 2', 'Zip Top 1', 'Zip Top 2', 'Zip Ground', 'rotate to course 1'
        ]

        if "Course Support 1" in role_to_column:
            course_roles.insert(0, "Course Support 1")  

        # Assign initial workers to course roles (9:00-9:30)
        course_workers = [valid_roles.get(role) for role in course_roles]

        # Rotate course roles for each subsequent time slot until lunch
        for slot_row in lunch_slots:
            # Rotate workers: Last worker moves to the first position
            course_workers = course_workers[-1:] + course_workers[:-1]

            # Assign rotated workers to their roles for this time slot
            for role, worker in zip(course_roles, course_workers):
                column = role_to_column.get(role)
                if column:
                    sheet.cell(row=slot_row, column=column).value = worker

        # Track assigned workers in the morning
        assigned_workers = set(valid_roles.values())  
        unassigned_workers = [worker.name for worker in in_today_workers + late_shift_workers if worker.name not in assigned_workers]

        if unassigned_workers:
            logging.warning(f"Unassigned workers found in the morning: {', '.join(unassigned_workers)}")

        # Track afternoon usage
        afternoon_valid_roles = {}  # Roles assigned in the afternoon
        afternoon_used_workers = set()  # Workers used in the afternoon

        # Define role categories for clarity and maintainability
        shed_roles = {'Host', 'Dekit', 'Kit Up 1', 'Kit Up 2', 'Kit Up 3', 'Clip In 1', 'Clip In 2'}
        tree_trek_roles = {'TREE TREK 1', 'TREE TREK 2'}
        course_roles = ['Course Support 2', 'Zip Top 1', 'Zip Top 2', 'Zip Ground', 'rotate to course 1']

        # Only add 'Course Support 1' if it exists in the Excel file
        if "Course Support 1" in role_to_column:
            course_roles.insert(0, "Course Support 1")

        # Convert course_roles to a **list** for ordering
        course_roles = list(course_roles)

        mini_trek_roles = {'Mini Trek'}
        ica_roles = [f"ICA {i}" for i in range(1, ica_morning_count + 1)]

        # Function to get eligible workers for afternoon assignments
        def get_afternoon_eligible_workers(role):
            eligible = [
                worker for worker in (in_today_workers + late_shift_workers)
                if worker.name not in afternoon_used_workers
            ]

            # Prefer different people for shed roles in the afternoon
            if role in shed_roles:
                # Try to avoid reusing the same person who did this role in the morning
                preferred = [
                    worker for worker in eligible
                    if role not in morning_assignments.get(worker.name, [])
                ]
                if preferred:
                    eligible = preferred
                else:
                    # Allow reuse only as fallback
                    logging.warning(f"⚠️ No new workers for {role}, reusing someone from the morning.")

            if role in role_to_training['KITUP']:
                eligible = [
                    worker for worker in eligible
                    if 'KITUP' in worker.roles  # Must be trained in KITUP
                ]

            elif role in tree_trek_roles:
                eligible = [
                    worker for worker in eligible
                    if 'AATT' in worker.roles  # Must be trained in AATT
                    and not any(m_role in tree_trek_roles for m_role in morning_assignments.get(worker.name, []))
                ]

            elif role in course_roles:
                if role == "Course Support 1" and "Course Support 1" not in role_to_column:
                    return []
                eligible = [
                    worker for worker in eligible
                    if 'AATT' in worker.roles  # Must be trained in AATT
                    and not any(m_role in course_roles for m_role in morning_assignments.get(worker.name, []))
                ]

            elif role in mini_trek_roles:
                eligible = [
                    worker for worker in eligible
                    if 'MT' in worker.roles  # Must be trained in Mini Trek
                    and not any(m_role in mini_trek_roles for m_role in morning_assignments.get(worker.name, []))
                ]

            elif role in ica_roles_afternoon:
                eligible = [
                    worker for worker in in_today_workers  # Ensure only early workers can be assigned
                    if worker.name not in afternoon_used_workers
                    and 'ICA' in worker.roles  # Must be trained in ICA
                    and not any(m_role in ica_roles for m_role in morning_assignments.get(worker.name, []))
                ]

            elif role in ["Course Support 2", "Zip Top 1", "Zip Top 2", "Zip Ground"]:
                # Ensure late workers are NOT assigned in these positions
                eligible = [
                    worker for worker in in_today_workers  # ONLY early workers
                    if worker.name not in afternoon_used_workers
                    and 'AATT' in worker.roles  # Must be trained in AATT
                ]

            return eligible
            
        ica_roles_afternoon = [f"ICA {i}" for i in range(1, ica_afternoon_count + 1)]

        prioritized_roles_afternoon = (
            ica_roles_afternoon +
            [
                'Mini Trek', 'Course Support 2', 'Zip Top 1', 'Zip Top 2', 'Zip Ground', 'rotate to course 1',
                'TREE TREK 1', 'TREE TREK 2',
                'Clip In 1', 'Clip In 2', 'Kit Up 3', 'Kit Up 2', 'Kit Up 1',
            ]
        )

        # Only add 'Course Support 1' if it exists in the Excel file
        if "Course Support 1" in role_to_column and "Course Support 1" not in prioritized_roles_afternoon:
            prioritized_roles_afternoon.insert(2, "Course Support 1")  # Put it with the other course roles

        

        # Assign workers for 12:45-1:30
        for role in prioritized_roles_afternoon:
            eligible_workers = get_afternoon_eligible_workers(role)
            if eligible_workers:
                selected_worker = choice(eligible_workers)  # Randomly select a worker
                afternoon_valid_roles[role] = selected_worker.name
                afternoon_used_workers.add(selected_worker.name)

        # Ensure unassigned KITUP-trained workers are placed in Kit Up and Clip In roles
        unassigned_kitup_workers_afternoon = [
            worker for worker in in_today_workers if worker.name not in afternoon_used_workers and 'KITUP' in worker.roles
        ]

        for role in kitup_roles_priority + kitup_roles_secondary:
            if role in role_to_column and role not in afternoon_valid_roles and unassigned_kitup_workers_afternoon:
                selected_worker = unassigned_kitup_workers_afternoon.pop(0)  # Assign first available KITUP worker
                afternoon_valid_roles[role] = selected_worker.name
                afternoon_used_workers.add(selected_worker.name)
                logging.debug(f"Assigning {selected_worker.name} to {role} (Afternoon KITUP role)")

        # Ensure all Kit Up roles are filled FIRST
        unassigned_kitup_workers_afternoon = [
            worker for worker in in_today_workers if worker.name not in afternoon_used_workers and 'KITUP' in worker.roles
        ]

        for role in kitup_roles_priority + kitup_roles_secondary:
            if role in role_to_column and role not in afternoon_valid_roles and unassigned_kitup_workers_afternoon:
                selected_worker = unassigned_kitup_workers_afternoon.pop(0)  # Assign first available KITUP worker
                afternoon_valid_roles[role] = selected_worker.name
                afternoon_used_workers.add(selected_worker.name)
                logging.debug(f"Assigning {selected_worker.name} to {role} (Afternoon KITUP role)")

        # Now, Assign Host & Dekit AFTER all Kit Up roles are filled
        unassigned_workers_afternoon = [
            worker for worker in in_today_workers if worker.name not in afternoon_used_workers
        ]

        assigned_host_dekit = set()  # Track assigned workers for Host & Dekit

        for role in ["Host", "Dekit"]:
            if role in role_to_column and role not in afternoon_valid_roles:  # Only assign if still empty
                available_untrained = [
                    worker for worker in unassigned_workers_afternoon
                    if worker.name not in assigned_host_dekit  # Ensure a different worker is assigned
                ]

                if available_untrained:
                    selected_worker = choice(available_untrained)
                    afternoon_valid_roles[role] = selected_worker.name  # Assign worker
                    afternoon_used_workers.add(selected_worker.name)  # Mark them as used
                    assigned_host_dekit.add(selected_worker.name)  # Track to avoid duplicate assignment

                    logging.debug(f"Assigning {selected_worker.name} to {role} (Afternoon Untrained Worker)")

        # Ensure Host & Dekit are printed for all afternoon time slots (12:45 PM - 4:00 PM)
        for slot_row in range(10, 16):  # Rows for 12:45, 1:30, ..., 4:00
            for role in ["Host", "Dekit"]:
                column = role_to_column.get(role)
                if column:
                    sheet.cell(row=slot_row, column=column).value = afternoon_valid_roles.get(role, "")

            # Write afternoon assignments (12:45-1:30) to the Excel sheet
            logging.info("Writing afternoon assignments (12:45-1:30) to the Excel sheet...")

            # Ensure the row for 12:45-1:30 is correctly found
            afternoon_slot_row = None

            for row in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value and str(cell_value).strip() == "12:45-1:30":
                    afternoon_slot_row = row
                    break

            # Ensure the assigned workers at 12:45 - 1:30 are **written** into the Excel sheet
            if afternoon_slot_row:
                for role, worker in afternoon_valid_roles.items():
                    column = role_to_column.get(role)
                    if column:
                        sheet.cell(row=afternoon_slot_row, column=column).value = worker if worker else ""

            # # Assign initial course and tree trek workers for the afternoon (12:45-1:30)
            # course_workers = [afternoon_valid_roles.get(role) for role in course_roles]
            tree_trek_workers = [afternoon_valid_roles.get(role) for role in tree_trek_roles]

            # Store afternoon workers before rotation ####
            saved_afternoon_workers = afternoon_valid_roles.copy() #####
            tree_trek_workers = [saved_afternoon_workers.get(role) for role in tree_trek_roles]

            # Ensure unassigned positions stay empty
            course_workers = [worker if worker else None for worker in course_workers]
            tree_trek_workers = [worker if worker else None for worker in tree_trek_workers]

            # Define afternoon time slots in the Excel sheet based on actual labels
            afternoon_slots_rows = [11, 12, 13, 14, 15]  # Corresponding rows for 13:30, 14:00, ..., 15:30

            # Rotate roles for each subsequent time slot in the afternoon
            for slot_index, slot_row in enumerate(afternoon_slots_rows):
                if slot_index > 0:
                    course_workers = course_workers[-1:] + course_workers[:-1]  # rotate the list

                for role, worker in zip(course_roles, course_workers):
                    column = role_to_column.get(role)
                    if column:
                        sheet.cell(row=slot_row, column=column).value = worker or ""

                # Assign rotated workers to course roles
                for role, worker in zip(course_roles, course_workers):
                    column = role_to_column.get(role)
                    if column:
                        sheet.cell(row=slot_row, column=column).value = worker if worker else ""

                # Assign rotated workers to tree trek roles
                for role, worker in zip(tree_trek_roles, tree_trek_workers):
                    column = role_to_column.get(role)
                    if column:
                        sheet.cell(row=slot_row, column=column).value = worker if worker else ""

            # Identify workers who were not assigned in the afternoon
            unassigned_workers = [worker.name for worker in in_today_workers + late_shift_workers if worker.name not in afternoon_valid_roles.values()]

            # Attempt to assign unassigned workers to open slots
            for role in afternoon_valid_roles:
                if not afternoon_valid_roles[role] and unassigned_workers:
                    afternoon_valid_roles[role] = unassigned_workers.pop(0)  # Assign first unassigned worker

            logging.info("\n======= AFTERNOON ASSIGNMENTS CHECK =======")
            for role, worker in afternoon_valid_roles.items():
                logging.info(f"{role} -> {worker}")
            logging.info("===========================================")

            # Ensure ICA workers assigned at 12:45 - 1:30 are stored for reuse
            ica_workers_after_lunch = {
                role: afternoon_valid_roles[role] for role in ica_roles if role in afternoon_valid_roles
            }

        # Assign workers for each afternoon time slot
        for slot_index, slot_row in enumerate(afternoon_slots_rows):

            # Assign ICA roles using the same workers from 12:45 - 1:30
            for role in ica_roles_afternoon:
                column = role_to_column.get(role)
                if column:
                    sheet.cell(row=slot_row, column=column).value = ica_workers_after_lunch.get(role, "")

            # Skip reassigning course roles that were already rotated and written
            prioritized_roles_afternoon = [
                role for role in prioritized_roles_afternoon
                if role not in course_roles and role not in ica_roles
            ]

            for role in prioritized_roles_afternoon:
                # ✅ Skip if already assigned at 12:45–1:30
                if role in afternoon_valid_roles:
                    continue
                if role in ica_roles:  # Skip ICA roles since they are already assigned
                    continue
                if role == "Mini Trek":  # ⛔️ Also skip re-assigning Mini Trek after 12:45–1:30
                    continue


                eligible_workers = get_afternoon_eligible_workers(role)

                if eligible_workers:
                    selected_worker = choice(eligible_workers)  # Randomly select a worker
                    afternoon_valid_roles[role] = selected_worker.name
                    afternoon_used_workers.add(selected_worker.name)

            # Handle Kit Up 2, Kit Up 3, Clip In 1, and Clip In 2 at 14:30
            if slot_row == 13:  # Row corresponding to 14:30
                # Swap Kit Up 2 with Clip In 1
                temp_kit_up_2 = afternoon_valid_roles.get('Kit Up 2')
                temp_kit_up_3 = afternoon_valid_roles.get('Kit Up 3')
                afternoon_valid_roles['Kit Up 2'], afternoon_valid_roles['Clip In 1'] = afternoon_valid_roles.get('Clip In 1'), temp_kit_up_2
                afternoon_valid_roles['Kit Up 3'], afternoon_valid_roles['Clip In 2'] = afternoon_valid_roles.get('Clip In 2'), temp_kit_up_3

            # Assign workers for the current time slot
            for role in prioritized_roles_afternoon:
                if role == "Mini Trek" and role in afternoon_valid_roles:
                    continue

                eligible_workers = get_afternoon_eligible_workers(role)
                
                # Filter out workers not trained for ICA roles if assigning to ICA roles
                if role.startswith('ICA'):
                    eligible_workers = [worker for worker in eligible_workers if 'ICA' in worker.roles]
                
                # Apply KITUP restrictions directly in assignment loop
                if role in ['Kit Up 1', 'Kit Up 2', 'Kit Up 3', 'Clip In 1', 'Clip In 2']:
                    eligible_workers = [worker for worker in eligible_workers if 'KITUP' in worker.roles]

                if eligible_workers:
                    selected_worker = choice(eligible_workers)  # Randomly select a worker
                    afternoon_valid_roles[role] = selected_worker.name
                    afternoon_used_workers.add(selected_worker.name)

            # Write assignments for the current time slot to the Excel sheet
            logging.info(f"Writing assignments for Row {slot_row} to the Excel sheet...")
            for role, worker in afternoon_valid_roles.items():
                if role == "Course Support 1" and "Course Support 1" not in role_to_column:
                    continue  # Skip if 'Course Support 1' is not in the Excel file

                column = role_to_column.get(role)
                if column:
                    sheet.cell(row=slot_row, column=column).value = worker

            # Handle Course role rotations for every half-hour time slot
            course_roles = ['Course Support 2', 'Zip Top 1', 'Zip Top 2', 'Zip Ground', 'rotate to course 1']

            # Only include 'Course Support 1' if it's in the Excel file
            if "Course Support 1" in role_to_column:
                course_roles.insert(0, "Course Support 1")

            if slot_index == 0:  # Initialize course roles for the first slot
                course_workers = [afternoon_valid_roles.get(role) for role in course_roles]

            # Rotate course roles immediately after the first slot
            if slot_index >= 0:
                course_workers = course_workers[-1:] + course_workers[:-1]  # Rotate the roles
                for role, worker in zip(course_roles, course_workers):
                    column = role_to_column.get(role)
                    if column:
                        sheet.cell(row=slot_row, column=column).value = worker

        evening_slots_rows = [16, 17, 18, 19, 20, 21]  # Rows for 16:00, 16:30, 17:00, 17:30, 18:00, 18:30

        # Ensure we only assign available late workers (if fewer than 4 exist)
        late_workers_for_ica = late_shift_workers[:min(4, len(late_shift_workers))]

        # Assign them for each evening time slot
        for slot_row in evening_slots_rows:
            for i, ica_role in enumerate(['ICA 1', 'ICA 2', 'ICA 3', 'ICA 4']):
                if i < len(late_workers_for_ica):  # Ensure we have a worker
                    worker = late_workers_for_ica[i]
                    column = role_to_column.get(ica_role)
                    if column:
                        sheet.cell(row=slot_row, column=column).value = worker.name
        
        # extend non-ICA roles printing up to selected cutoff hour
        # row map: 10..15 are 12:45,13:30,14:00,14:30,15:00,15:30
        # row map: 16..21 are 16:00,16:30,17:00,17:30,18:00,18:30
        # for a cutoff of 16 print up to 15:30
        # for a cutoff of 17 print up to 16:30
        # for a cutoff of 18 print up to 17:30

        cutoff_rows_map = {
            16: [15],                 # up to 15:30
            17: [15, 16, 17],         # up to 16:30
            18: [15, 16, 17, 18, 19], # up to 17:30
        }
        non_ica_rows_to_fill = cutoff_rows_map.get(print_until_hour, [])

        # ordered list of course roles for evening reuse
        course_roles_evening = ['Course Support 2', 'Zip Top 1', 'Zip Top 2', 'Zip Ground', 'rotate to course 1']
        if "Course Support 1" in role_to_column:
            course_roles_evening.insert(0, "Course Support 1")

        # seed the rotation list from the latest known order
        if not course_workers:
            course_workers = [afternoon_valid_roles.get(role) for role in course_roles_evening]

        saved_afternoon_workers_all = afternoon_valid_roles.copy()

        for slot_row in non_ica_rows_to_fill:
            if slot_row >= 16 and course_workers:
                course_workers = course_workers[-1:] + course_workers[:-1]

            for role, worker in zip(course_roles_evening, course_workers):
                col = role_to_column.get(role)
                if col:
                    sheet.cell(row=slot_row, column=col).value = worker or ""

            for role in ['TREE TREK 1', 'TREE TREK 2']:
                col = role_to_column.get(role)
                if col:
                    sheet.cell(row=slot_row, column=col).value = saved_afternoon_workers_all.get(role, "")

            for role in ['Host', 'Dekit', 'Kit Up 1', 'Kit Up 2', 'Kit Up 3', 'Clip In 1', 'Clip In 2', 'Mini Trek']:
                col = role_to_column.get(role)
                if col:
                    sheet.cell(row=slot_row, column=col).value = saved_afternoon_workers_all.get(role, "")


        logging.info("\n======== SPARE WORKERS SUMMARY ========")
        logging.info(f"Morning Spare Workers ({len(morning_spare_workers)}): {', '.join(morning_spare_workers) if morning_spare_workers else 'None'}")
        afternoon_spare_workers = [
            worker.name for worker in in_today_workers + late_shift_workers
            if worker.name not in afternoon_valid_roles.values()
        ]

        # Log afternoon spare workers clearly
        if afternoon_spare_workers:
            logging.info(f"Afternoon Spare Workers ({len(afternoon_spare_workers)}): {', '.join(afternoon_spare_workers)}")
        else:
            logging.info("No Afternoon Spare Workers found.")

        logging.info("========================================")

        # Fallback: Force assign Host and Dekit in afternoon if still unassigned and workers are left
        for role in ["Host", "Dekit"]:
            if role in role_to_column and role not in afternoon_valid_roles:
                available_spares = [
                    worker for worker in in_today_workers + late_shift_workers
                    if worker.name not in afternoon_used_workers
                ]

                if available_spares:
                    selected_worker = choice(available_spares)
                    afternoon_valid_roles[role] = selected_worker.name
                    afternoon_used_workers.add(selected_worker.name)
                    logging.warning(f"⚠️ Fallback assigning {selected_worker.name} to {role} (afternoon fallback).")


        # Track assigned workers in the afternoon
        assigned_workers = set(afternoon_valid_roles.values())  
        unassigned_workers = [worker.name for worker in in_today_workers + late_shift_workers if worker.name not in assigned_workers]

        if unassigned_workers:
            logging.warning(f"Unassigned workers found in the afternoon: {', '.join(unassigned_workers)}")
        
        # spare summary section (after planner, around line 24+)
        summary_start_row = 24

        header_cell = sheet.cell(row=summary_start_row - 1, column=1)
        header_cell.value = "Spare"
        header_cell.font = Font(bold=True, size=14, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Morning Spare Summary
        morning_summary = (
            f"Morning Spare Workers: {', '.join(morning_spare_workers)}"
            if morning_spare_workers
            else "Morning Spare Workers: No spare"
        )
        sheet.cell(row=summary_start_row, column=1).value = morning_summary

        # Afternoon Spare Summary
        afternoon_summary = (
            f"Afternoon Spare Workers: {', '.join(afternoon_spare_workers)}"
            if afternoon_spare_workers
            else "Afternoon Spare Workers: No spare"
        )
        sheet.cell(row=summary_start_row + 1, column=1).value = afternoon_summary

        # Define where to write the "Workers In Today" summary
        summary_col = 27  # Column T
        summary_start_row = 1

        # Header
        header_cell = sheet.cell(row=summary_start_row, column=summary_col)
        header_cell.value = "Instructors"
        header_cell.font = Font(bold=True, size=12, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")

        # Sort workers by actual start time
        in_today_sorted = []

        for worker in in_today_workers + late_shift_workers:
            for availability in worker.availability:
                try:
                    start = parser.parse(availability["start"])
                    end = parser.parse(availability["end"])

                    # Normalize to local tz for display/sorting
                    if start.tzinfo is None:
                        start_local = start.replace(tzinfo=TIMEZONE)
                    else:
                        start_local = start.astimezone(TIMEZONE)

                    if end.tzinfo is None:
                        end_local = end.replace(tzinfo=TIMEZONE)
                    else:
                        end_local = end.astimezone(TIMEZONE)

                    # Check date against the selected date (also in local tz)
                    if start_local.date() <= selected_date.astimezone(TIMEZONE).date() <= end_local.date():
                        in_today_sorted.append((start_local, worker.name, start_local, end_local))
                        break
                except Exception as e:
                    logging.warning(f"Skipping invalid availability for {worker.name}: {e}")

        # Sort by start time
        in_today_sorted.sort(key=lambda x: x[0])

        # Write each worker
        for i, (_, name, start_local, end_local) in enumerate(in_today_sorted, start=1):
            row = summary_start_row + i
            time_range = f"{start_local.strftime('%H:%M')} - {end_local.strftime('%H:%M')}"
            sheet.cell(row=row, column=summary_col).value = f"{name} - {time_range}"

        # Save and send the Excel file
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="day_schedule.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        logging.error(f"Error generating schedule: {e}")
        return jsonify({'error': str(e)}), 500


# API endpoint to create a worker
@app.route("/workers", methods=["POST"])
def create_worker():
    try:
        logging.debug("Incoming request data: %s", request.json)
        data = request.get_json()

        if not data or "name" not in data or "roles" not in data or "availability" not in data:
            return jsonify({"error": "Missing required fields: name, roles, or availability"}), 400

        new_worker = Worker(
            name=data["name"],
            roles=data["roles"],
            availability=data["availability"],
        )
        db.session.add(new_worker)
        db.session.commit()

        logging.info(f"Worker {new_worker.name} created successfully.")
        return jsonify({
            "message": "Worker created successfully",
            "worker": {
                "id": new_worker.id,
                "name": new_worker.name,
                "roles": new_worker.roles,
                "availability": new_worker.availability,
            },
        }), 201
    except Exception as e:
        logging.error(f"Error creating worker: {e}")
        return jsonify({"error": str(e)}), 500

# API endpoint to update a worker by ID
@app.route("/workers/<int:worker_id>", methods=["PUT"])
def update_worker(worker_id):
    try:
        worker = Worker.query.get(worker_id)
        if not worker:
            return jsonify({"error": f"No worker found with ID {worker_id}"}), 404

        data = request.get_json()

        worker.name = data.get("name", worker.name)
        worker.roles = data.get("roles", worker.roles)

        if "availability" in data:
            worker.availability = [
                {
                    "start": parser.parse(a["start"]).isoformat(),
                    "end": parser.parse(a["end"]).isoformat(),
                    "late": bool(a.get("late", False)),
                }
                for a in data["availability"]
            ]

        db.session.commit()

        return jsonify({
            "message": "Worker updated successfully",
            "worker": {
                "id": worker.id,
                "name": worker.name,
                "roles": worker.roles,
                "availability": worker.availability,
            },
        }), 200
    except Exception as e:
        logging.error(f"Error updating worker {worker_id}: {e}")
        return jsonify({"error": str(e)}), 500


# API endpoint to delete a worker by ID
@app.route("/workers/<int:worker_id>", methods=["DELETE"])
def delete_worker(worker_id):
    try:
        worker = Worker.query.get(worker_id)
        if not worker:
            return jsonify({"error": f"Worker with ID {worker_id} not found"}), 404

        db.session.delete(worker)
        db.session.commit()

        return jsonify({"message": "Worker deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/login", methods=["POST"])
def login():
    data = request.get_json() or {}
    password = data.get("password", "")
    expected = os.getenv("ADMIN_PASSWORD", "CenterParcs")
    if compare_digest(password, expected):
        return jsonify({"success": True}), 200
    return jsonify({"success": False, "error": "Invalid password"}), 401


# Home route to confirm the app is running
@app.route("/")
def home():
    return "Flask app is running!"

# Create the database tables
with app.app_context():
    db.create_all()
    logging.info("Database tables created successfully!")

# Run the app
if __name__ == "__main__":
    logging.info("Starting Flask app...")
    app.run(debug=True, port=5001)